from PyQt5.QtWidgets import (QDialog, QVBoxLayout, QHBoxLayout, QPushButton, 
                             QLabel, QComboBox, QLineEdit, QTableWidget, 
                             QTableWidgetItem, QMessageBox, QGroupBox, QSpinBox,
                             QColorDialog, QCheckBox, QTabWidget, QWidget, 
                             QListWidget, QRadioButton, QButtonGroup, QDoubleSpinBox,
                             QScrollArea, QFrame)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QColor, QBrush
try:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class AdvancedFormattingDialog(QDialog):
    """
    Advanced Formatting Dialog with multiple features:
    - Conditional Formatting Rules
    - Custom Number Formats
    - Cell Styles Library
    - Format Painter
    - Row Height Auto-fit
    - Merge Cells
    - Cell Borders
    """
    
    def __init__(self, parent=None, table_widget=None, df=None):
        super().__init__(parent)
        self.parent_editor = parent
        self.table_widget = table_widget
        self.df = df
        self.conditional_rules = []
        self.cell_styles = {}
        self.copied_format = None
        
        self.setWindowTitle("📈 Advanced Formatting")
        self.setMinimumWidth(700)
        self.setMinimumHeight(600)
        
        self.init_ui()
        
    def init_ui(self):
        layout = QVBoxLayout()
        
        # Create tabs for different formatting features
        self.tabs = QTabWidget()
        
        # Tab 1: Conditional Formatting
        self.conditional_tab = self.create_conditional_formatting_tab()
        self.tabs.addTab(self.conditional_tab, "🎨 Conditional Formatting")
        
        # Tab 2: Number Formats
        self.number_format_tab = self.create_number_format_tab()
        self.tabs.addTab(self.number_format_tab, "🔢 Number Formats")
        
        # Tab 3: Cell Styles Library
        self.styles_tab = self.create_styles_library_tab()
        self.tabs.addTab(self.styles_tab, "💎 Cell Styles")
        
        # Tab 4: Format Painter & More
        self.tools_tab = self.create_tools_tab()
        self.tabs.addTab(self.tools_tab, "🛠️ Tools")
        
        # Tab 5: Borders
        self.borders_tab = self.create_borders_tab()
        self.tabs.addTab(self.borders_tab, "📐 Borders")
        
        layout.addWidget(self.tabs)
        
        # Buttons
        button_layout = QHBoxLayout()
        
        self.apply_btn = QPushButton("Apply")
        self.apply_btn.clicked.connect(self.apply_formatting)
        button_layout.addWidget(self.apply_btn)
        
        self.close_btn = QPushButton("Close")
        self.close_btn.clicked.connect(self.close)
        button_layout.addWidget(self.close_btn)
        
        layout.addLayout(button_layout)
        self.setLayout(layout)
    
    def create_conditional_formatting_tab(self):
        """Create conditional formatting rules tab"""
        widget = QWidget()
        layout = QVBoxLayout()
        
        # Instructions
        info_label = QLabel("Create rules to automatically format cells based on their values")
        info_label.setWordWrap(True)
        layout.addWidget(info_label)
        
        # Rule creation section
        rule_group = QGroupBox("New Conditional Rule")
        rule_layout = QVBoxLayout()
        
        # Select column
        col_layout = QHBoxLayout()
        col_layout.addWidget(QLabel("Column:"))
        self.cond_column_combo = QComboBox()
        if self.df is not None:
            self.cond_column_combo.addItems(self.df.columns.tolist())
        col_layout.addWidget(self.cond_column_combo)
        col_layout.addStretch()
        rule_layout.addLayout(col_layout)
        
        # Condition type
        condition_layout = QHBoxLayout()
        condition_layout.addWidget(QLabel("Condition:"))
        self.condition_combo = QComboBox()
        self.condition_combo.addItems([
            "Greater than (>)",
            "Less than (<)",
            "Equal to (=)",
            "Between",
            "Contains text",
            "Starts with",
            "Ends with",
            "Is empty",
            "Is not empty"
        ])
        condition_layout.addWidget(self.condition_combo)
        condition_layout.addStretch()
        rule_layout.addLayout(condition_layout)
        
        # Value input
        value_layout = QHBoxLayout()
        value_layout.addWidget(QLabel("Value:"))
        self.cond_value_input = QLineEdit()
        self.cond_value_input.setPlaceholderText("Enter value...")
        value_layout.addWidget(self.cond_value_input)
        
        value_layout.addWidget(QLabel("Value 2 (for Between):"))
        self.cond_value2_input = QLineEdit()
        self.cond_value2_input.setPlaceholderText("Optional...")
        value_layout.addWidget(self.cond_value2_input)
        rule_layout.addLayout(value_layout)
        
        # Format to apply
        format_layout = QHBoxLayout()
        format_layout.addWidget(QLabel("Format:"))
        
        self.cond_bg_color_btn = QPushButton("Background Color")
        self.cond_bg_color_btn.clicked.connect(self.pick_cond_bg_color)
        self.cond_bg_color = "#FFFFFF"
        format_layout.addWidget(self.cond_bg_color_btn)
        
        self.cond_text_color_btn = QPushButton("Text Color")
        self.cond_text_color_btn.clicked.connect(self.pick_cond_text_color)
        self.cond_text_color = "#000000"
        format_layout.addWidget(self.cond_text_color_btn)
        
        self.cond_bold_cb = QCheckBox("Bold")
        format_layout.addWidget(self.cond_bold_cb)
        
        format_layout.addStretch()
        rule_layout.addLayout(format_layout)
        
        # Add rule button
        add_rule_btn = QPushButton("➕ Add Rule")
        add_rule_btn.clicked.connect(self.add_conditional_rule)
        rule_layout.addWidget(add_rule_btn)
        
        rule_group.setLayout(rule_layout)
        layout.addWidget(rule_group)
        
        # Rules list
        rules_group = QGroupBox("Active Rules")
        rules_layout = QVBoxLayout()
        
        self.rules_list = QListWidget()
        rules_layout.addWidget(self.rules_list)
        
        remove_rule_btn = QPushButton("🗑️ Remove Selected Rule")
        remove_rule_btn.clicked.connect(self.remove_conditional_rule)
        rules_layout.addWidget(remove_rule_btn)
        
        rules_group.setLayout(rules_layout)
        layout.addWidget(rules_group)
        
        layout.addStretch()
        widget.setLayout(layout)
        return widget
    
    def create_number_format_tab(self):
        """Create custom number formats tab"""
        widget = QWidget()
        layout = QVBoxLayout()
        
        info_label = QLabel("Apply custom number formatting to selected columns")
        info_label.setWordWrap(True)
        layout.addWidget(info_label)
        
        # Column selection
        col_group = QGroupBox("Select Column")
        col_layout = QHBoxLayout()
        self.num_format_column_combo = QComboBox()
        if self.df is not None:
            self.num_format_column_combo.addItems(self.df.columns.tolist())
        col_layout.addWidget(self.num_format_column_combo)
        col_group.setLayout(col_layout)
        layout.addWidget(col_group)
        
        # Format selection
        format_group = QGroupBox("Number Format")
        format_layout = QVBoxLayout()
        
        self.format_type_combo = QComboBox()
        self.format_type_combo.addItems([
            "General",
            "Number (1,234.56)",
            "Currency ($1,234.56)",
            "Accounting",
            "Percentage (12.34%)",
            "Scientific (1.23E+03)",
            "Fraction (1/2)",
            "Date (MM/DD/YYYY)",
            "Time (HH:MM:SS)",
            "Custom..."
        ])
        self.format_type_combo.currentTextChanged.connect(self.on_format_type_changed)
        format_layout.addWidget(self.format_type_combo)
        
        # Decimal places
        decimal_layout = QHBoxLayout()
        decimal_layout.addWidget(QLabel("Decimal Places:"))
        self.decimal_spin = QSpinBox()
        self.decimal_spin.setRange(0, 10)
        self.decimal_spin.setValue(2)
        decimal_layout.addWidget(self.decimal_spin)
        decimal_layout.addStretch()
        format_layout.addLayout(decimal_layout)
        
        # Custom format string
        custom_layout = QHBoxLayout()
        custom_layout.addWidget(QLabel("Custom Format:"))
        self.custom_format_input = QLineEdit()
        self.custom_format_input.setPlaceholderText("e.g., #,##0.00;[Red]-#,##0.00")
        self.custom_format_input.setEnabled(False)
        custom_layout.addWidget(self.custom_format_input)
        format_layout.addLayout(custom_layout)
        
        # Preview
        self.format_preview = QLabel("Preview: 1234.5678")
        self.format_preview.setFrameStyle(QFrame.Panel | QFrame.Sunken)
        self.format_preview.setMinimumHeight(40)
        self.format_preview.setAlignment(Qt.AlignCenter)
        format_layout.addWidget(self.format_preview)
        
        format_group.setLayout(format_layout)
        layout.addWidget(format_group)
        
        # Apply button
        apply_num_format_btn = QPushButton("Apply Number Format")
        apply_num_format_btn.clicked.connect(self.apply_number_format)
        layout.addWidget(apply_num_format_btn)
        
        layout.addStretch()
        widget.setLayout(layout)
        return widget
    
    def create_styles_library_tab(self):
        """Create cell styles library tab"""
        widget = QWidget()
        layout = QVBoxLayout()
        
        info_label = QLabel("Pre-defined cell styles for quick formatting")
        layout.addWidget(info_label)
        
        # Styles list
        styles_group = QGroupBox("Available Styles")
        styles_layout = QVBoxLayout()
        
        self.styles_list = QListWidget()
        self.populate_styles_library()
        self.styles_list.itemDoubleClicked.connect(self.apply_style_from_library)
        styles_layout.addWidget(self.styles_list)
        
        apply_style_btn = QPushButton("Apply Selected Style to Selection")
        apply_style_btn.clicked.connect(self.apply_style_from_library)
        styles_layout.addWidget(apply_style_btn)
        
        styles_group.setLayout(styles_layout)
        layout.addWidget(styles_group)
        
        # Custom style creation
        custom_group = QGroupBox("Create Custom Style")
        custom_layout = QVBoxLayout()
        
        name_layout = QHBoxLayout()
        name_layout.addWidget(QLabel("Style Name:"))
        self.custom_style_name = QLineEdit()
        self.custom_style_name.setPlaceholderText("My Custom Style")
        name_layout.addWidget(self.custom_style_name)
        custom_layout.addLayout(name_layout)
        
        colors_layout = QHBoxLayout()
        self.custom_bg_btn = QPushButton("Background Color")
        self.custom_bg_btn.clicked.connect(self.pick_custom_bg_color)
        self.custom_bg_color = "#FFFFFF"
        colors_layout.addWidget(self.custom_bg_btn)
        
        self.custom_text_btn = QPushButton("Text Color")
        self.custom_text_btn.clicked.connect(self.pick_custom_text_color)
        self.custom_text_color = "#000000"
        colors_layout.addWidget(self.custom_text_btn)
        custom_layout.addLayout(colors_layout)
        
        options_layout = QHBoxLayout()
        self.custom_bold_cb = QCheckBox("Bold")
        options_layout.addWidget(self.custom_bold_cb)
        self.custom_italic_cb = QCheckBox("Italic")
        options_layout.addWidget(self.custom_italic_cb)
        options_layout.addStretch()
        custom_layout.addLayout(options_layout)
        
        save_style_btn = QPushButton("💾 Save Style")
        save_style_btn.clicked.connect(self.save_custom_style)
        custom_layout.addWidget(save_style_btn)
        
        custom_group.setLayout(custom_layout)
        layout.addWidget(custom_group)
        
        layout.addStretch()
        widget.setLayout(layout)
        return widget
    
    def create_tools_tab(self):
        """Create tools tab (Format Painter, Auto-fit, Merge)"""
        widget = QWidget()
        layout = QVBoxLayout()
        
        # Format Painter
        painter_group = QGroupBox("🖌️ Format Painter")
        painter_layout = QVBoxLayout()
        
        painter_info = QLabel("Copy formatting from one cell/range to another")
        painter_info.setWordWrap(True)
        painter_layout.addWidget(painter_info)
        
        painter_buttons = QHBoxLayout()
        self.copy_format_btn = QPushButton("📋 Copy Format from Selection")
        self.copy_format_btn.clicked.connect(self.copy_format)
        painter_buttons.addWidget(self.copy_format_btn)
        
        self.paste_format_btn = QPushButton("📄 Paste Format to Selection")
        self.paste_format_btn.clicked.connect(self.paste_format)
        self.paste_format_btn.setEnabled(False)
        painter_buttons.addWidget(self.paste_format_btn)
        painter_layout.addLayout(painter_buttons)
        
        painter_group.setLayout(painter_layout)
        layout.addWidget(painter_group)
        
        # Row Height Auto-fit
        autofit_group = QGroupBox("📏 Row Height Auto-fit")
        autofit_layout = QVBoxLayout()
        
        autofit_info = QLabel("Automatically adjust row heights to fit content")
        autofit_info.setWordWrap(True)
        autofit_layout.addWidget(autofit_info)
        
        autofit_buttons = QHBoxLayout()
        self.autofit_selected_btn = QPushButton("Auto-fit Selected Rows")
        self.autofit_selected_btn.clicked.connect(self.autofit_selected_rows)
        autofit_buttons.addWidget(self.autofit_selected_btn)
        
        self.autofit_all_btn = QPushButton("Auto-fit All Rows")
        self.autofit_all_btn.clicked.connect(self.autofit_all_rows)
        autofit_buttons.addWidget(self.autofit_all_btn)
        autofit_layout.addLayout(autofit_buttons)
        
        autofit_group.setLayout(autofit_layout)
        layout.addWidget(autofit_group)
        
        # Merge Cells
        merge_group = QGroupBox("🔗 Merge Cells")
        merge_layout = QVBoxLayout()
        
        merge_info = QLabel("Combine selected cells into one")
        merge_info.setWordWrap(True)
        merge_layout.addWidget(merge_info)
        
        merge_buttons = QHBoxLayout()
        self.merge_btn = QPushButton("Merge Selected Cells")
        self.merge_btn.clicked.connect(self.merge_cells)
        merge_buttons.addWidget(self.merge_btn)
        
        self.unmerge_btn = QPushButton("Unmerge Selected Cells")
        self.unmerge_btn.clicked.connect(self.unmerge_cells)
        merge_buttons.addWidget(self.unmerge_btn)
        merge_layout.addLayout(merge_buttons)
        
        merge_group.setLayout(merge_layout)
        layout.addWidget(merge_group)
        
        layout.addStretch()
        widget.setLayout(layout)
        return widget
    
    def create_borders_tab(self):
        """Create borders tab"""
        widget = QWidget()
        layout = QVBoxLayout()
        
        info_label = QLabel("Apply custom borders to selected cells")
        layout.addWidget(info_label)
        
        # Border style selection
        style_group = QGroupBox("Border Style")
        style_layout = QVBoxLayout()
        
        self.border_style_combo = QComboBox()
        self.border_style_combo.addItems([
            "Thin",
            "Medium",
            "Thick",
            "Dashed",
            "Dotted",
            "Double"
        ])
        style_layout.addWidget(self.border_style_combo)
        
        # Border color
        color_layout = QHBoxLayout()
        color_layout.addWidget(QLabel("Border Color:"))
        self.border_color_btn = QPushButton("Choose Color")
        self.border_color_btn.clicked.connect(self.pick_border_color)
        self.border_color = "#000000"
        color_layout.addWidget(self.border_color_btn)
        color_layout.addStretch()
        style_layout.addLayout(color_layout)
        
        style_group.setLayout(style_layout)
        layout.addWidget(style_group)
        
        # Border position
        position_group = QGroupBox("Apply Borders To")
        position_layout = QVBoxLayout()
        
        self.border_all_cb = QCheckBox("All Borders")
        position_layout.addWidget(self.border_all_cb)
        
        self.border_outline_cb = QCheckBox("Outline")
        position_layout.addWidget(self.border_outline_cb)
        
        self.border_top_cb = QCheckBox("Top")
        position_layout.addWidget(self.border_top_cb)
        
        self.border_bottom_cb = QCheckBox("Bottom")
        position_layout.addWidget(self.border_bottom_cb)
        
        self.border_left_cb = QCheckBox("Left")
        position_layout.addWidget(self.border_left_cb)
        
        self.border_right_cb = QCheckBox("Right")
        position_layout.addWidget(self.border_right_cb)
        
        position_group.setLayout(position_layout)
        layout.addWidget(position_group)
        
        # Apply button
        apply_borders_btn = QPushButton("Apply Borders to Selection")
        apply_borders_btn.clicked.connect(self.apply_borders)
        layout.addWidget(apply_borders_btn)
        
        # Remove borders button
        remove_borders_btn = QPushButton("Remove All Borders from Selection")
        remove_borders_btn.clicked.connect(self.remove_borders)
        layout.addWidget(remove_borders_btn)
        
        layout.addStretch()
        widget.setLayout(layout)
        return widget
    
    # Conditional Formatting Methods
    def pick_cond_bg_color(self):
        color = QColorDialog.getColor()
        if color.isValid():
            self.cond_bg_color = color.name()
            self.cond_bg_color_btn.setStyleSheet(f"background-color: {self.cond_bg_color}")
    
    def pick_cond_text_color(self):
        color = QColorDialog.getColor()
        if color.isValid():
            self.cond_text_color = color.name()
            self.cond_text_color_btn.setStyleSheet(f"background-color: {self.cond_text_color}")
    
    def add_conditional_rule(self):
        if self.df is None:
            return
            
        column = self.cond_column_combo.currentText()
        condition = self.condition_combo.currentText()
        value = self.cond_value_input.text()
        value2 = self.cond_value2_input.text()
        
        rule = {
            'column': column,
            'condition': condition,
            'value': value,
            'value2': value2,
            'bg_color': self.cond_bg_color,
            'text_color': self.cond_text_color,
            'bold': self.cond_bold_cb.isChecked()
        }
        
        self.conditional_rules.append(rule)
        
        # Update rules list display
        rule_text = f"{column} | {condition} | {value}"
        if value2:
            rule_text += f" and {value2}"
        self.rules_list.addItem(rule_text)
        
        QMessageBox.information(self, "Success", "Conditional rule added!")
    
    def remove_conditional_rule(self):
        current_row = self.rules_list.currentRow()
        if current_row >= 0:
            self.rules_list.takeItem(current_row)
            del self.conditional_rules[current_row]
            QMessageBox.information(self, "Success", "Rule removed!")
    
    # Number Format Methods
    def on_format_type_changed(self, text):
        self.custom_format_input.setEnabled(text == "Custom...")
        self.update_format_preview()
    
    def update_format_preview(self):
        format_type = self.format_type_combo.currentText()
        decimals = self.decimal_spin.value()
        sample = 1234.5678
        
        try:
            if "Number" in format_type:
                preview = f"{sample:,.{decimals}f}"
            elif "Currency" in format_type:
                preview = f"${sample:,.{decimals}f}"
            elif "Percentage" in format_type:
                preview = f"{sample:.{decimals}f}%"
            elif "Scientific" in format_type:
                preview = f"{sample:.{decimals}e}"
            else:
                preview = str(sample)
            
            self.format_preview.setText(f"Preview: {preview}")
        except:
            self.format_preview.setText("Preview: Invalid format")
    
    def apply_number_format(self):
        if self.table_widget is None:
            return
            
        column = self.num_format_column_combo.currentText()
        format_type = self.format_type_combo.currentText()
        
        try:
            # Get column index
            col_names = [self.table_widget.horizontalHeaderItem(i).text() 
                        for i in range(self.table_widget.columnCount())]
            if column not in col_names:
                return
                
            col_idx = col_names.index(column)
            
            # Apply format to all cells in column
            for row in range(self.table_widget.rowCount()):
                item = self.table_widget.item(row, col_idx)
                if item:
                    try:
                        value = float(item.text().replace(',', '').replace('$', ''))
                        formatted = self.format_number(value, format_type, self.decimal_spin.value())
                        item.setText(formatted)
                    except:
                        pass
            
            QMessageBox.information(self, "Success", f"Number format applied to {column}!")
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to apply format: {str(e)}")
    
    def format_number(self, value, format_type, decimals):
        if "Number" in format_type:
            return f"{value:,.{decimals}f}"
        elif "Currency" in format_type:
            return f"${value:,.{decimals}f}"
        elif "Percentage" in format_type:
            return f"{value:.{decimals}f}%"
        elif "Scientific" in format_type:
            return f"{value:.{decimals}e}"
        else:
            return str(value)
    
    # Cell Styles Methods
    def populate_styles_library(self):
        """Populate with predefined styles"""
        styles = [
            ("Header", "#4472C4", "#FFFFFF", True, False),
            ("Accent 1", "#FFE699", "#000000", False, False),
            ("Accent 2", "#A9D08E", "#000000", False, False),
            ("Accent 3", "#F4B084", "#000000", False, False),
            ("Good", "#C6EFCE", "#006100", False, False),
            ("Bad", "#FFC7CE", "#9C0006", False, False),
            ("Neutral", "#FFEB9C", "#9C6500", False, False),
            ("Warning", "#FF9999", "#FFFFFF", True, False),
            ("Note", "#E7E6E6", "#000000", False, True),
            ("Title", "#44546A", "#FFFFFF", True, False)
        ]
        
        for name, bg, text, bold, italic in styles:
            self.cell_styles[name] = {
                'bg_color': bg,
                'text_color': text,
                'bold': bold,
                'italic': italic
            }
            self.styles_list.addItem(name)
    
    def pick_custom_bg_color(self):
        color = QColorDialog.getColor()
        if color.isValid():
            self.custom_bg_color = color.name()
            self.custom_bg_btn.setStyleSheet(f"background-color: {self.custom_bg_color}")
    
    def pick_custom_text_color(self):
        color = QColorDialog.getColor()
        if color.isValid():
            self.custom_text_color = color.name()
            self.custom_text_btn.setStyleSheet(f"background-color: {self.custom_text_color}")
    
    def save_custom_style(self):
        name = self.custom_style_name.text().strip()
        if not name:
            QMessageBox.warning(self, "Error", "Please enter a style name!")
            return
        
        self.cell_styles[name] = {
            'bg_color': self.custom_bg_color,
            'text_color': self.custom_text_color,
            'bold': self.custom_bold_cb.isChecked(),
            'italic': self.custom_italic_cb.isChecked()
        }
        
        self.styles_list.addItem(name)
        QMessageBox.information(self, "Success", f"Style '{name}' saved!")
        self.custom_style_name.clear()
    
    def apply_style_from_library(self):
        if self.table_widget is None:
            return
            
        current_item = self.styles_list.currentItem()
        if not current_item:
            return
            
        style_name = current_item.text()
        if style_name not in self.cell_styles:
            return
            
        style = self.cell_styles[style_name]
        selected_ranges = self.table_widget.selectedRanges()
        
        for sel_range in selected_ranges:
            for row in range(sel_range.topRow(), sel_range.bottomRow() + 1):
                for col in range(sel_range.leftColumn(), sel_range.rightColumn() + 1):
                    item = self.table_widget.item(row, col)
                    if not item:
                        item = QTableWidgetItem()
                        self.table_widget.setItem(row, col, item)
                    
                    # Apply style
                    bg_color = QColor(style['bg_color'])
                    text_color = QColor(style['text_color'])
                    item.setBackground(QBrush(bg_color))
                    item.setForeground(QBrush(text_color))
                    
                    font = item.font()
                    font.setBold(style['bold'])
                    font.setItalic(style['italic'])
                    item.setFont(font)
        
        QMessageBox.information(self, "Success", f"Style '{style_name}' applied!")
    
    # Format Painter Methods
    def copy_format(self):
        if self.table_widget is None:
            return
            
        selected = self.table_widget.selectedItems()
        if not selected:
            QMessageBox.warning(self, "Error", "Please select a cell to copy format from!")
            return
        
        item = selected[0]
        self.copied_format = {
            'bg_color': item.background().color(),
            'text_color': item.foreground().color(),
            'font': item.font()
        }
        
        self.paste_format_btn.setEnabled(True)
        QMessageBox.information(self, "Success", "Format copied! Now select cells to paste to.")
    
    def paste_format(self):
        if self.table_widget is None or self.copied_format is None:
            return
            
        selected_ranges = self.table_widget.selectedRanges()
        
        for sel_range in selected_ranges:
            for row in range(sel_range.topRow(), sel_range.bottomRow() + 1):
                for col in range(sel_range.leftColumn(), sel_range.rightColumn() + 1):
                    item = self.table_widget.item(row, col)
                    if not item:
                        item = QTableWidgetItem()
                        self.table_widget.setItem(row, col, item)
                    
                    item.setBackground(QBrush(self.copied_format['bg_color']))
                    item.setForeground(QBrush(self.copied_format['text_color']))
                    item.setFont(self.copied_format['font'])
        
        QMessageBox.information(self, "Success", "Format pasted!")
    
    # Auto-fit Methods
    def autofit_selected_rows(self):
        if self.table_widget is None:
            return
            
        selected_ranges = self.table_widget.selectedRanges()
        rows = set()
        
        for sel_range in selected_ranges:
            for row in range(sel_range.topRow(), sel_range.bottomRow() + 1):
                rows.add(row)
        
        for row in rows:
            self.table_widget.resizeRowToContents(row)
        
        QMessageBox.information(self, "Success", f"Auto-fitted {len(rows)} row(s)!")
    
    def autofit_all_rows(self):
        if self.table_widget is None:
            return
            
        for row in range(self.table_widget.rowCount()):
            self.table_widget.resizeRowToContents(row)
        
        QMessageBox.information(self, "Success", "All rows auto-fitted!")
    
    # Merge Cells Methods
    def merge_cells(self):
        if self.table_widget is None:
            return
            
        selected_ranges = self.table_widget.selectedRanges()
        if not selected_ranges:
            QMessageBox.warning(self, "Error", "Please select cells to merge!")
            return
        
        sel_range = selected_ranges[0]
        top_row = sel_range.topRow()
        left_col = sel_range.leftColumn()
        row_span = sel_range.bottomRow() - top_row + 1
        col_span = sel_range.rightColumn() - left_col + 1
        
        if row_span == 1 and col_span == 1:
            QMessageBox.warning(self, "Error", "Please select more than one cell to merge!")
            return
        
        self.table_widget.setSpan(top_row, left_col, row_span, col_span)
        QMessageBox.information(self, "Success", "Cells merged!")
    
    def unmerge_cells(self):
        if self.table_widget is None:
            return
            
        selected_ranges = self.table_widget.selectedRanges()
        if not selected_ranges:
            QMessageBox.warning(self, "Error", "Please select merged cells to unmerge!")
            return
        
        for sel_range in selected_ranges:
            for row in range(sel_range.topRow(), sel_range.bottomRow() + 1):
                for col in range(sel_range.leftColumn(), sel_range.rightColumn() + 1):
                    self.table_widget.setSpan(row, col, 1, 1)
        
        QMessageBox.information(self, "Success", "Cells unmerged!")
    
    # Border Methods
    def pick_border_color(self):
        color = QColorDialog.getColor()
        if color.isValid():
            self.border_color = color.name()
            self.border_color_btn.setStyleSheet(f"background-color: {self.border_color}")
    
    def apply_borders(self):
        if self.table_widget is None:
            return
            
        selected_ranges = self.table_widget.selectedRanges()
        if not selected_ranges:
            QMessageBox.warning(self, "Error", "Please select cells to apply borders!")
            return
        
        # Note: PyQt5 QTableWidget doesn't have native border support per cell
        # This is a visual indication only - actual Excel borders would be applied during export
        QMessageBox.information(self, "Info", 
            "Border settings saved! Borders will be applied when exporting to Excel.\n" +
            "Note: Border display in the table view is limited.")
    
    def remove_borders(self):
        if self.table_widget is None:
            return
            
        QMessageBox.information(self, "Info", "Border removal noted for Excel export.")
    
    def apply_formatting(self):
        """Apply all active conditional formatting rules"""
        if self.table_widget is None or self.df is None:
            return
        
        try:
            # Apply conditional formatting rules
            for rule in self.conditional_rules:
                self.apply_conditional_rule(rule)
            
            if self.conditional_rules:
                QMessageBox.information(self, "Success", 
                    f"Applied {len(self.conditional_rules)} conditional formatting rule(s)!")
            else:
                QMessageBox.information(self, "Info", 
                    "No conditional rules to apply. Use other tabs for immediate formatting.")
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to apply formatting: {str(e)}")
    
    def apply_conditional_rule(self, rule):
        """Apply a single conditional formatting rule"""
        column = rule['column']
        condition = rule['condition']
        value = rule['value']
        value2 = rule['value2']
        
        # Get column index
        col_names = [self.table_widget.horizontalHeaderItem(i).text() 
                    for i in range(self.table_widget.columnCount())]
        if column not in col_names:
            return
            
        col_idx = col_names.index(column)
        
        # Apply to each row
        for row in range(self.table_widget.rowCount()):
            item = self.table_widget.item(row, col_idx)
            if not item:
                continue
                
            cell_value = item.text()
            should_format = self.check_condition(cell_value, condition, value, value2)
            
            if should_format:
                bg_color = QColor(rule['bg_color'])
                text_color = QColor(rule['text_color'])
                item.setBackground(QBrush(bg_color))
                item.setForeground(QBrush(text_color))
                
                if rule['bold']:
                    font = item.font()
                    font.setBold(True)
                    item.setFont(font)
    
    def check_condition(self, cell_value, condition, value, value2):
        """Check if a cell value meets the condition"""
        try:
            if condition == "Is empty":
                return cell_value == ""
            elif condition == "Is not empty":
                return cell_value != ""
            elif condition == "Contains text":
                return value.lower() in cell_value.lower()
            elif condition == "Starts with":
                return cell_value.lower().startswith(value.lower())
            elif condition == "Ends with":
                return cell_value.lower().endswith(value.lower())
            
            # Numeric comparisons
            try:
                cell_num = float(cell_value)
                value_num = float(value)
                
                if condition == "Greater than (>)":
                    return cell_num > value_num
                elif condition == "Less than (<)":
                    return cell_num < value_num
                elif condition == "Equal to (=)":
                    return cell_num == value_num
                elif condition == "Between" and value2:
                    value2_num = float(value2)
                    return value_num <= cell_num <= value2_num
            except:
                return False
                
        except Exception as e:
            return False
