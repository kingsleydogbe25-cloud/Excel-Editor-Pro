"""
FormatApplier: Helper module to apply formatting settings to Excel files
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class FormatApplier:
    """Apply formatting settings to Excel workbooks"""
    
    @staticmethod
    def apply_formatting(worksheet, format_settings, dataframe):
        """
        Apply formatting settings to a worksheet
        
        Args:
            worksheet: openpyxl worksheet object
            format_settings: dictionary of column formatting settings
            dataframe: pandas DataFrame with the data
        """
        if not format_settings:
            return
        
        # Get column name to letter mapping
        col_mapping = {str(col): get_column_letter(idx + 1) 
                      for idx, col in enumerate(dataframe.columns)}
        
        for col_name, settings in format_settings.items():
            if col_name not in col_mapping:
                continue
            
            col_letter = col_mapping[col_name]
            
            # Apply column width
            if settings.get('width') == 'auto':
                worksheet.column_dimensions[col_letter].auto_size = True
            else:
                # Convert pixels to Excel width units (approximate)
                excel_width = settings.get('width', 100) / 7
                worksheet.column_dimensions[col_letter].width = excel_width
            
            # Apply formatting to all cells in the column
            for row_idx in range(2, len(dataframe) + 2):  # Start from row 2 (after header)
                cell = worksheet[f'{col_letter}{row_idx}']
                
                # Alignment
                h_align = settings.get('h_align', 'general')
                v_align = settings.get('v_align', 'center')
                wrap_text = settings.get('wrap_text', False)
                
                cell.alignment = Alignment(
                    horizontal=h_align,
                    vertical=v_align,
                    wrap_text=wrap_text
                )
                
                # Font
                font_name = settings.get('font_name', 'Calibri')
                font_size = settings.get('font_size', 11)
                bold = settings.get('bold', False)
                italic = settings.get('italic', False)
                underline = 'single' if settings.get('underline', False) else None
                text_color = settings.get('text_color', '#000000').lstrip('#')
                
                cell.font = Font(
                    name=font_name,
                    size=font_size,
                    bold=bold,
                    italic=italic,
                    underline=underline,
                    color=text_color
                )
                
                # Background color
                bg_color = settings.get('bg_color', '#FFFFFF').lstrip('#')
                if bg_color and bg_color != 'FFFFFF':
                    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
                
                # Number format
                cell.number_format = FormatApplier._get_number_format(settings)
            
            # Apply header formatting if specified
            if settings.get('format_header', True):
                header_cell = worksheet[f'{col_letter}1']
                
                # Header font
                header_bold = settings.get('header_bold', True)
                header_text_color = settings.get('header_text_color', '#FFFFFF').lstrip('#')
                
                header_cell.font = Font(
                    name=settings.get('font_name', 'Calibri'),
                    size=settings.get('font_size', 11),
                    bold=header_bold,
                    color=header_text_color
                )
                
                # Header background
                header_bg = settings.get('header_bg_color', '#4472C4').lstrip('#')
                header_cell.fill = PatternFill(start_color=header_bg, end_color=header_bg, fill_type='solid')
                
                # Header alignment
                header_cell.alignment = Alignment(
                    horizontal='center',
                    vertical='center',
                    wrap_text=True
                )
    
    @staticmethod
    def _get_number_format(settings):
        """Generate Excel number format string from settings"""
        format_type = settings.get('format_type', 'General')
        decimal_places = settings.get('decimal_places', 2)
        use_thousands = settings.get('use_thousands', False)
        currency_symbol = settings.get('currency_symbol', '$')
        
        if format_type == 'General':
            return 'General'
        
        elif format_type == 'Number':
            dec_format = '0' * decimal_places
            if use_thousands:
                return f'#,##0.{dec_format}' if decimal_places > 0 else '#,##0'
            else:
                return f'0.{dec_format}' if decimal_places > 0 else '0'
        
        elif format_type == 'Currency':
            dec_format = '0' * decimal_places
            if use_thousands:
                return f'{currency_symbol}#,##0.{dec_format}' if decimal_places > 0 else f'{currency_symbol}#,##0'
            else:
                return f'{currency_symbol}0.{dec_format}' if decimal_places > 0 else f'{currency_symbol}0'
        
        elif format_type == 'Accounting':
            dec_format = '0' * decimal_places
            return f'_({currency_symbol}* #,##0.{dec_format}_);_({currency_symbol}* (#,##0.{dec_format});_({currency_symbol}* "-"??_);_(@_)'
        
        elif format_type == 'Percentage':
            dec_format = '0' * decimal_places
            return f'0.{dec_format}%' if decimal_places > 0 else '0%'
        
        elif format_type == 'Date':
            return 'mm/dd/yyyy'
        
        elif format_type == 'Time':
            return 'h:mm:ss AM/PM'
        
        elif format_type == 'Scientific':
            dec_format = '0' * decimal_places
            return f'0.{dec_format}E+00' if decimal_places > 0 else '0E+00'
        
        elif format_type == 'Text':
            return '@'
        
        return 'General'
    
    @staticmethod
    def apply_borders(worksheet, dataframe):
        """Apply borders to all cells with data"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(1, len(dataframe) + 2):  # Include header
            for col in range(1, len(dataframe.columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = thin_border
